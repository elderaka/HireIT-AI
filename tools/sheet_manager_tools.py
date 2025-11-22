
"""
Sheet Manager Tools for watsonx Orchestrate ADK.

Import with:
  orchestrate tools import -k python -f tools/sheet_manager_tools.py

Core idea:
- Work with PUBLIC Google Drive / Sheets links.
- Prefer Sheets export CSV when it's a native Google Sheet.
- Fallback to Drive download when export is blocked or file is uploaded XLSX.

Tools:
- extract_drive_file_id(link)
- make_sheets_export_link(file_id, format="csv"|"xlsx", gid=0)
- make_drive_download_link(file_id)
- download_public_bytes(url)
- sniff_html_interstitial(file_bytes)
- sniff_table_file_type(file_bytes)
- parse_csv_bytes(file_bytes)
- parse_xlsx_bytes(file_bytes)
- apply_sheet_patch(rows, patch_spec)
- build_csv_bytes(rows, headers=None)
- parse_sheet_public(link_or_id, gid=0, prefer="csv")
"""

from ibm_watsonx_orchestrate.agent_builder.tools import tool
from typing import Dict, Any, Optional, List, Union
import re
import csv
from io import StringIO, BytesIO

try:
    import requests  # type: ignore
except Exception:
    requests = None  # type: ignore


# ---- helpers ----

_PATTERNS = [
    re.compile(r"drive\.google\.com/file/d/([a-zA-Z0-9_-]+)"),
    re.compile(r"docs\.google\.com/(?:document|spreadsheets|presentation)/d/([a-zA-Z0-9_-]+)"),
    re.compile(r"drive\.google\.com/uc\?[^#]*\bid=([a-zA-Z0-9_-]+)"),
    re.compile(r"/d/([a-zA-Z0-9_-]+)"),
    re.compile(r"[?&]id=([a-zA-Z0-9_-]+)"),
]

def _extract_id(link: str) -> Optional[str]:
    if not link:
        return None
    s = link.strip()
    for pat in _PATTERNS:
        m = pat.search(s)
        if m:
            return m.group(1)
    return None

def _is_html_response(file_bytes: bytes, content_type: Optional[str]) -> bool:
    if content_type and "text/html" in content_type.lower():
        return True
    head = (file_bytes or b"")[:300].lstrip()
    try:
        head_txt = head.decode("utf-8", errors="replace").lower()
    except Exception:
        head_txt = ""
    return head_txt.startswith("<!doctype html") or "<html" in head_txt or "google drive" in head_txt


# ---- tools ----

@tool
def extract_drive_file_id(link: str) -> Dict[str, Any]:
    """Extract a Google Drive/Docs/Sheets file_id from a link."""
    return {"file_id": _extract_id(link)}

@tool
def make_sheets_export_link(file_id: str, format: str = "csv", gid: int = 0) -> Dict[str, Any]:
    """Make a public export URL for a native Google Sheet."""
    fmt = (format or "csv").lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "csv"
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format={fmt}&gid={gid}"
    return {"export_url": url}

@tool
def make_drive_download_link(file_id: str) -> Dict[str, Any]:
    """Make a public Drive download URL for any Drive file (including uploaded XLSX)."""
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    return {"drive_url": url}

@tool
def download_public_bytes(url: str, timeout_sec: int = 30) -> Dict[str, Any]:
    """Download bytes from a public URL. Returns http_code/content_type/file_bytes."""
    if requests is None:
        return {"http_code": 0, "content_type": None, "file_bytes": b""}
    r = requests.get(url, timeout=timeout_sec)
    return {
        "http_code": int(r.status_code),
        "content_type": r.headers.get("Content-Type"),
        "file_bytes": r.content or b""
    }

@tool
def sniff_html_interstitial(file_bytes: bytes = b"", b: Optional[bytes] = None) -> Dict[str, Any]:
    """Detect HTML interstitial; accepts file_bytes or legacy b."""
    if (not file_bytes) and b:
        file_bytes = b
    head = (file_bytes or b"")[:300].lstrip()
    try:
        head_txt = head.decode("utf-8", errors="replace").lower()
    except Exception:
        head_txt = ""
    is_html = head_txt.startswith("<!doctype html") or "<html" in head_txt or "google drive" in head_txt
    return {"html": bool(is_html), "head": head_txt[:200]}

@tool
def sniff_table_file_type(file_bytes: bytes) -> Dict[str, Any]:
    """Infer file type: csv vs xlsx vs txt using magic bytes + content sniff."""
    b0 = (file_bytes or b"")[:8]
    if b0.startswith(b"PK\x03\x04"):
        return {"file_type": "xlsx"}
    try:
        sample = file_bytes[:2000].decode("utf-8", errors="replace")
    except Exception:
        sample = ""
    if "," in sample or "\n" in sample:
        return {"file_type": "csv"}
    return {"file_type": "txt"}

@tool
def parse_csv_bytes(file_bytes: bytes, encoding: str = "utf-8", limit_rows: int = 5000) -> Dict[str, Any]:
    """Parse CSV bytes into headers/table/rows."""
    warnings: List[str] = []
    text = file_bytes.decode(encoding, errors="replace")
    reader = csv.reader(StringIO(text))
    table: List[List[str]] = []
    for i, row in enumerate(reader):
        if i >= limit_rows:
            warnings.append(f"Row limit {limit_rows} reached; truncated.")
            break
        table.append(row)

    headers = table[0] if table else []
    rows = []
    for r in table[1:]:
        obj = {}
        for j, h in enumerate(headers):
            obj[h] = r[j] if j < len(r) else ""
        rows.append(obj)

    return {
        "file_type": "csv",
        "headers": headers,
        "table": table,
        "rows": rows,
        "meta": {"row_count": len(rows), "warnings": warnings}
    }

@tool
def parse_xlsx_bytes(file_bytes: bytes, limit_rows: int = 2000) -> Dict[str, Any]:
    """Parse XLSX bytes if openpyxl is available; else warn."""
    warnings: List[str] = []
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        warnings.append(f"openpyxl unavailable: {e}")
        return {"file_type": "xlsx", "sheets": [], "meta": {"warnings": warnings}}

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheets_out = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheet_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= limit_rows:
                warnings.append(f"{name}: row limit {limit_rows} reached; truncated.")
                break
            sheet_rows.append([cell for cell in row])
        sheets_out.append({"sheet": name, "rows": sheet_rows})

    return {"file_type": "xlsx", "sheets": sheets_out, "meta": {"warnings": warnings}}

@tool
def apply_sheet_patch(rows: List[Dict[str, Any]], patch_spec: Any) -> Dict[str, Any]:
    """Apply patch_spec over rows (update/append/delete)."""
    if rows is None:
        rows = []
    patches = patch_spec if isinstance(patch_spec, list) else [patch_spec]
    changes = 0
    updated = list(rows)

    def match_where(row: Dict[str, Any], where: Dict[str, Any]) -> bool:
        for k, v in (where or {}).items():
            if str(row.get(k, "")) != str(v):
                return False
        return True

    for p in patches:
        if not isinstance(p, dict):
            continue
        op = p.get("op")
        if op == "update":
            where = p.get("where", {})
            sets = p.get("set", {})
            for row in updated:
                if match_where(row, where):
                    for k, v in sets.items():
                        row[k] = v
                    changes += 1
        elif op == "append":
            row = p.get("row")
            if isinstance(row, dict):
                updated.append(row)
                changes += 1
        elif op == "delete":
            where = p.get("where", {})
            before = len(updated)
            updated = [r for r in updated if not match_where(r, where)]
            changes += (before - len(updated))

    return {"rows": updated, "meta": {"changes": changes}}

@tool
def build_csv_bytes(rows: List[Dict[str, Any]], headers: Optional[List[str]] = None, encoding: str = "utf-8") -> Dict[str, Any]:
    """Build CSV bytes from rows + headers."""
    if rows is None:
        rows = []
    if headers is None:
        headers = []
        seen = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    headers.append(k)
                    seen.add(k)

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h, "") for h in headers])

    bts = out.getvalue().encode(encoding)
    return {"file_bytes": bts, "headers": headers}

@tool
def parse_sheet_public(
    link_or_id: str,
    gid: int = 0,
    prefer: str = "csv",
    timeout_sec: int = 30
) -> Dict[str, Any]:
    """
    One-shot public Sheets/Drive reader.

    Steps:
    1) get file_id from link_or_id if needed
    2) try Sheets export (native Sheets) -> download bytes
    3) if html / non-200 -> fallback Drive download
    4) sniff file type then parse accordingly
    """
    file_id = link_or_id if re.fullmatch(r"[A-Za-z0-9_-]{10,}", link_or_id or "") else _extract_id(link_or_id)
    if not file_id:
        return {"ok": False, "error": "Could not extract file_id", "file_id": None}

    # try export first
    export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format={prefer}&gid={gid}"
    exp = download_public_bytes(export_url, timeout_sec=timeout_sec)
    html_flag = _is_html_response(exp.get("file_bytes", b""), exp.get("content_type"))

    file_bytes = exp.get("file_bytes", b"")
    source = "sheets_export"

    if exp.get("http_code") != 200 or html_flag:
        # fallback to drive
        drive_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        drv = download_public_bytes(drive_url, timeout_sec=timeout_sec)
        file_bytes = drv.get("file_bytes", b"")
        source = "drive_download"

    sniff = sniff_table_file_type(file_bytes)
    ftype = sniff.get("file_type", "txt")

    if ftype == "csv":
        parsed = parse_csv_bytes(file_bytes)
    elif ftype == "xlsx":
        parsed = parse_xlsx_bytes(file_bytes)
    else:
        parsed = {"file_type": ftype, "text": file_bytes[:2000].decode("utf-8", errors="replace")}

    parsed["ok"] = True
    parsed["file_id"] = file_id
    parsed["source"] = source
    return parsed
