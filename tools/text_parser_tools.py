"""
Text Parser Tools for watsonx Orchestrate ADK.

Import with:
  orchestrate tools import -k python -f tools/text_parser_tools.py

Tools:
- parse_drive_public_link(link, ext=None, encoding="utf-8"):
    Extracts file_id from a PUBLIC Drive link, downloads bytes, auto-sniffs type if ext is None.
- parse_file_bytes(file_bytes, file_name=None, ext=None, encoding="utf-8"):
    Parses raw bytes directly, auto-sniffs type if no ext/file_name.

Notes:
- Drive downloading only works if the file is shared as
  "Anyone with the link can view/download".
"""

from ibm_watsonx_orchestrate.agent_builder.tools import tool
from typing import Optional, Any, Dict, List
import json
import re
import zlib
from io import BytesIO

try:
    import requests  # type: ignore
except Exception:
    requests = None  # type: ignore

import zipfile


# ---- helpers ----

_PATTERNS = [
    re.compile(r"drive\.google\.com/file/d/([a-zA-Z0-9_-]+)"),
    re.compile(r"docs\.google\.com/(?:document|spreadsheets|presentation)/d/([a-zA-Z0-9_-]+)"),
    re.compile(r"drive\.google\.com/uc\?[^#]*\bid=([a-zA-Z0-9_-]+)"),
    re.compile(r"/d/([a-zA-Z0-9_-]+)"),
]

def _extract_drive_id(link: str) -> Optional[str]:
    if not link:
        return None
    s = link.strip()
    for pat in _PATTERNS:
        m = pat.search(s)
        if m:
            return m.group(1)
    if "id=" in s:
        m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", s)
        if m:
            return m.group(1)
    return None

def _infer_ext(file_name: Optional[str], ext: Optional[str]) -> Optional[str]:
    if ext:
        return ext.lower().lstrip(".")
    if file_name and "." in file_name:
        return file_name.rsplit(".", 1)[-1].lower()
    return None

def _safe_decode(b: bytes, encoding: str = "utf-8") -> str:
    try:
        return b.decode(encoding)
    except Exception:
        return b.decode("utf-8", errors="replace")

def _sniff_office_type(file_bytes: bytes) -> Optional[str]:
    # ZIP container sniff (xlsx/docx/pptx) without external deps
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as z:
            names = z.namelist()
            if any(n.startswith("xl/") for n in names):
                return "xlsx"
            if any(n.startswith("word/") for n in names):
                return "docx"
            if any(n.startswith("ppt/") for n in names):
                return "pptx"
    except Exception:
        pass
    return None

def _sniff_file_type(file_bytes: bytes, header_hint: Optional[str] = None) -> str:
    b0 = file_bytes[:8]

    # 1) Magic bytes
    if b0.startswith(b"%PDF-"):
        return "pdf"
    if b0.startswith(b"PK\x03\x04"):
        office = _sniff_office_type(file_bytes)
        if office:
            return office
        return "zip"

    stripped = file_bytes.lstrip()[:1]
    if stripped in (b"{", b"["):
        return "json"

    # 2) Content-Type hint
    if header_hint:
        ct = header_hint.lower()
        if "pdf" in ct:
            return "pdf"
        if "spreadsheet" in ct or "excel" in ct:
            return "xlsx"
        if "json" in ct:
            return "json"
        if "csv" in ct:
            return "csv"
        if "text" in ct:
            return "txt"

    return "txt"



def _pdf_unescape(s: bytes) -> str:
    out = bytearray()
    i = 0
    while i < len(s):
        c = s[i]
        if c == 0x5C:  # backslash
            i += 1
            if i >= len(s): break
            nxt = s[i]

            if nxt in b"nrtbf":
                out.extend({
                    ord("n"): b"\n",
                    ord("r"): b"\r",
                    ord("t"): b"\t",
                    ord("b"): b"\b",
                    ord("f"): b"\f",
                }[nxt])
                i += 1
                continue

            if nxt in b"()\\":
                out.append(nxt); i += 1; continue

            if 0x30 <= nxt <= 0x37:
                oct_digits = bytes([nxt]); i += 1
                for _ in range(2):
                    if i < len(s) and 0x30 <= s[i] <= 0x37:
                        oct_digits += bytes([s[i]]); i += 1
                    else:
                        break
                try: out.append(int(oct_digits, 8))
                except Exception: pass
                continue

            out.append(nxt); i += 1
        else:
            out.append(c); i += 1

    return out.decode("utf-8", errors="replace")

def _collapse_short_word_runs(text: str) -> str:
    """
    Collapse sequences of short alphabetic tokens (len<=3) into one token.
    """
    pat = re.compile(r'(?:(?:\b[A-Za-z]{1,3}\b\s+){2,}\b[A-Za-z]{1,3}\b)')
    def repl(m):
        return m.group(0).replace(" ", "")
    return pat.sub(repl, text)


def _despace_spans(text: str, span_pat: re.Pattern) -> str:
    """
    Find spans that match span_pat and remove spaces inside them.
    """
    def repl(m):
        s = m.group(0)
        return re.sub(r"\s+", "", s)
    return span_pat.sub(repl, text)


def _normalize_contacts(text: str) -> str:
    """
    Remove PDF kerning spaces specifically inside emails, urls, and phone numbers.
    """
    # Email-like span allowing spaces
    email_span = re.compile(
        r'(?i)[A-Z0-9._%+\-\s]+@\s*[A-Z0-9.\-\s]+\.\s*[A-Z\s]{2,}'
    )
    text = _despace_spans(text, email_span)

    # URL-like spans (http(s) or domainy things)
    url_span = re.compile(
        r'(?i)(https?\s*:\s*//\s*[^\s|]+|[A-Z0-9\-\s]+\.\s*com\s*/\s*[^\s|]+)'
    )
    text = _despace_spans(text, url_span)
    # also fix spaced "linkedin.com/in/..."
    text = re.sub(r'(?i)l\s*i\s*n\s*k\s*e\s*d\s*i\s*n', 'linkedin', text)

    # Phone span: +digits with spaces/hyphens
    phone_span = re.compile(r'\+?\d(?:[\s\-]?\d){6,}')
    text = _despace_spans(text, phone_span)

    return text

def _merge_pdf_chunks(chunks):
    out = []
    prev = ""

    for c in chunks:
        c = c.strip()
        if not c:
            continue

        if prev and len(prev) == 1 and len(c) == 1:
            out[-1] = out[-1] + c
            prev = out[-1]
            continue

        if prev and len(prev) == 1 and len(c) > 1:
            out[-1] = out[-1] + c
            prev = out[-1]
            continue

        if prev and not prev.endswith((" ", "\n")) and not c.startswith(("\n", " ")):
            out.append(" " + c)
            prev = c
        else:
            out.append(c)
            prev = c

    text = "".join(out)

    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    text = _collapse_short_word_runs(text)
    text = _normalize_contacts(text)

    return text.strip()



def _extract_text_from_pdf_fallback(file_bytes: bytes) -> str:
    """
    Fallback extractor that:
    1) Finds FlateDecode streams, decompresses them,
    2) Scans decompressed content for Tj/TJ text operators.
    Works for most text-based PDFs, including ATS CVs.
    """
    data = file_bytes

    # 1) Find all "stream ... endstream" blocks
    stream_pat = re.compile(rb"stream\s*(.*?)\s*endstream", re.DOTALL)
    # 2) Detect FlateDecode near the stream (simple heuristic)
    flate_hint_pat = re.compile(rb"/Filter\s*/FlateDecode")

    # Tj/TJ patterns
    tj_pat = re.compile(rb"\((?:\\.|[^\\)])*\)\s*Tj")
    tj_array_pat = re.compile(rb"\[(.*?)\]\s*TJ", re.DOTALL)
    inner_str_pat = re.compile(rb"\((?:\\.|[^\\)])*\)")

    chunks = []

    for m in stream_pat.finditer(data):
        raw_stream = m.group(1)

        # Look backwards a bit for /FlateDecode
        start = max(0, m.start() - 200)
        context = data[start:m.start()]
        is_flate = bool(flate_hint_pat.search(context))

        candidates = []

        if is_flate:
            # Try zlib decompression (some PDFs need "raw" mode)
            for wbits in (zlib.MAX_WBITS, -zlib.MAX_WBITS):
                try:
                    candidates.append(zlib.decompress(raw_stream, wbits))
                    break
                except Exception:
                    continue
        else:
            candidates.append(raw_stream)

        for cand in candidates:
            for tm in tj_pat.finditer(cand):
                sm = inner_str_pat.search(tm.group(0))
                if sm:
                    lit = sm.group(0)[1:-1]
                    chunks.append(_pdf_unescape(lit))

            for tm in tj_array_pat.finditer(cand):
                arr = tm.group(1)
                for sm in inner_str_pat.finditer(arr):
                    lit = sm.group(0)[1:-1]
                    chunks.append(_pdf_unescape(lit))

    return _merge_pdf_chunks(chunks)


def _parse_bytes(
    file_bytes: bytes,
    file_type: str,
    encoding: str,
    max_chars: int,
    meta: Dict[str, Any],
) -> Dict[str, Any]:
    text: str = ""
    tables: List[List[Any]] = []
    obj: Any = None

    if file_type in ("txt", "md"):
        text = _safe_decode(file_bytes, encoding)

    elif file_type == "csv":
        text = _safe_decode(file_bytes, encoding)
        rows = []
        for line in text.splitlines():
            rows.append([c.strip() for c in line.split(",")])
        tables = rows

    elif file_type == "json":
        raw = _safe_decode(file_bytes, encoding)
        try:
            obj = json.loads(raw)
            text = json.dumps(obj, ensure_ascii=False, indent=2)
        except Exception as e:
            meta["warnings"].append(f"Invalid JSON: {e}")
            text = raw

    elif file_type == "pdf":
        # Try proper parser first, fallback if unavailable
        try:
            import PyPDF2  # type: ignore
            reader = PyPDF2.PdfReader(BytesIO(file_bytes))
            parts = []
            for page in reader.pages:
                parts.append(page.extract_text() or "")
            text = "\n\n".join(parts)

        except Exception as e:
            # Fallback: crude PDF text extraction (works for ATS-friendly text PDFs)
            meta["warnings"].append(
                f"PyPDF2 unavailable or failed ({e}). Using fallback PDF text extractor."
            )
            text = _extract_text_from_pdf_fallback(file_bytes)
            if not text.strip():
                meta["warnings"].append(
                    "Fallback extractor found no text. PDF may be image-based or heavily encoded."
                )


    elif file_type in ("xlsx", "xls"):
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            sheets_out = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheet_rows = []
                for row in ws.iter_rows(values_only=True):
                    sheet_rows.append([cell for cell in row])
                sheets_out.append({"sheet": name, "rows": sheet_rows})
            obj = {"sheets": sheets_out}

            preview_lines = []
            for s in sheets_out:
                preview_lines.append(f"## {s['sheet']}")
                for r in s["rows"][:100]:
                    preview_lines.append(
                        "\t".join("" if v is None else str(v) for v in r)
                    )
            text = "\n".join(preview_lines)

        except Exception as e:
            meta["warnings"].append(
                f"Excel parsing failed. openpyxl may be unavailable: {e}"
            )
            text = ""

    else:
        meta["warnings"].append(
            f"Unsupported/unknown file_type '{file_type}'. Returning raw text decode."
        )
        text = _safe_decode(file_bytes, encoding)

    if len(text) > max_chars:
        meta["warnings"].append(f"Text truncated to {max_chars} chars.")
        text = text[:max_chars]

    return {
        "file_type": file_type,
        "text": text,
        "tables": tables,
        "obj": obj,
        "meta": meta,
    }


# ---- tools ----

@tool
def debug_head_bytes(file_bytes: bytes, n: int = 200) -> Dict[str, Any]:
    """
    Return first N bytes as latin-1 and hex to verify what we downloaded.
    """
    head = file_bytes[:n]
    return {
        "head_latin1": head.decode("latin-1", errors="replace"),
        "head_hex": head.hex()[:400]
    }

@tool
def parse_drive_public_link(
    link: str,
    ext: Optional[str] = None,
    encoding: str = "utf-8",
    max_chars: int = 200000,
    timeout_sec: int = 30,
    include_warnings: bool = False
) -> Dict[str, Any]:
    """
    Download a PUBLIC Google Drive file from its share link and parse it.

    - Extracts file_id from the link.
    - Downloads bytes via `https://drive.google.com/uc?export=download&id=...`
      (file must be shared to "Anyone with link").
    - Auto-sniffs file type if ext is not provided.
    - Returns structured JSON with file_type, file_name, text, tables, obj, meta.

    Set include_warnings=true only if you want diagnostic warnings back.
    """
    meta: Dict[str, Any] = {"warnings": []}

    file_id = _extract_drive_id(link)
    if not file_id:
        meta["warnings"].append("Could not extract file_id from link.")
        out = {"file_type": "unknown", "text": "", "tables": [], "obj": None, "meta": meta}
        if not include_warnings:
            out["meta"]["warnings"] = []
        out["file_name"] = None
        return out

    if requests is None:
        meta["warnings"].append("requests library unavailable; cannot download Drive file.")
        out = {"file_type": "unknown", "text": "", "tables": [], "obj": None, "meta": meta}
        if not include_warnings:
            out["meta"]["warnings"] = []
        out["file_name"] = None
        return out

    url = f"https://drive.google.com/uc?export=download&id={file_id}"

    try:
        r = requests.get(url, timeout=timeout_sec)
        if r.status_code != 200:
            meta["warnings"].append(
                f"Drive download failed (HTTP {r.status_code}). "
                "Check sharing: 'Anyone with the link can download'."
            )
            out = {"file_type": "unknown", "text": "", "tables": [], "obj": None, "meta": meta}
            if not include_warnings:
                out["meta"]["warnings"] = []
            out["file_name"] = None
            return out

        file_bytes = r.content
        header_ct = r.headers.get("Content-Type")
        sniffed = _sniff_file_type(file_bytes, header_ct)

        cd = r.headers.get("Content-Disposition", "")
        fname = None
        m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd)
        if m:
            fname = m.group(1)

    except Exception as e:
        meta["warnings"].append(f"Drive download error: {e}")
        out = {"file_type": "unknown", "text": "", "tables": [], "obj": None, "meta": meta}
        if not include_warnings:
            out["meta"]["warnings"] = []
        out["file_name"] = None
        return out

    ext_hint = _infer_ext(None, ext)
    if ext_hint and ext_hint != sniffed:
        meta["warnings"].append(
            f"File looks like '{sniffed}' but you forced ext='{ext_hint}'."
        )
        chosen_type = ext_hint
    else:
        chosen_type = ext_hint or sniffed

    parsed = _parse_bytes(file_bytes, chosen_type, encoding, max_chars, meta)

    if not fname:
        fname = f"drive_{file_id}.{chosen_type}"
    parsed["file_name"] = fname

    if not include_warnings:
        parsed["meta"]["warnings"] = []

    return parsed


@tool
def parse_file_bytes(
    file_bytes: bytes,
    file_name: Optional[str] = None,
    ext: Optional[str] = None,
    encoding: str = "utf-8",
    max_chars: int = 200000,
    include_warnings: bool = False
) -> Dict[str, Any]:
    """
    Parse raw file bytes into structured text/tables.

    - If ext/file_name is missing, auto-sniffs file type from magic bytes.
    - Supports txt/md/csv/json/pdf/xlsx.
    - Returns structured JSON with file_type, file_name, text, tables, obj, meta.

    Set include_warnings=true only if you want diagnostic warnings back.
    """
    meta: Dict[str, Any] = {"warnings": []}

    ext_hint = _infer_ext(file_name, ext)
    sniffed = _sniff_file_type(file_bytes)

    if ext_hint and ext_hint != sniffed:
        meta["warnings"].append(
            f"File looks like '{sniffed}' but ext/file_name suggests '{ext_hint}'."
        )
        chosen_type = ext_hint
    else:
        chosen_type = ext_hint or sniffed

    parsed = _parse_bytes(file_bytes, chosen_type, encoding, max_chars, meta)

    parsed["file_name"] = file_name or f"uploaded.{chosen_type}"

    if not include_warnings:
        parsed["meta"]["warnings"] = []

    return parsed